"""
Export Service - Generates PDF and Excel reports from project BOQ data.
"""
import io
from datetime import datetime
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


DARK_RED = colors.HexColor("#C0392B")
LIGHT_RED = colors.HexColor("#FADBD8")
MID_GRAY = colors.HexColor("#95A5A6")
DARK_GRAY = colors.HexColor("#2C3E50")
LIGHT_GRAY = colors.HexColor("#ECF0F1")
WHITE = colors.white


def generate_pdf(project: dict, building_data: dict, recommendations: dict, boq_sections: list) -> bytes:
    """Generate a professional PDF BOQ report."""
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title=f"Fire BOQ - {project.get('project_name', '')}",
    )

    styles = getSampleStyleSheet()
    story = []

    # ── HEADER ───────────────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        fontSize=18, textColor=DARK_RED, spaceAfter=4,
        alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=10, textColor=DARK_GRAY, spaceAfter=2,
        alignment=TA_CENTER
    )

    story.append(Paragraph("FIRE PROTECTION SYSTEM", title_style))
    story.append(Paragraph("BILL OF QUANTITIES (BOQ)", title_style))
    story.append(Spacer(1, 4))
    story.append(HRFlowable(width="100%", thickness=2, color=DARK_RED))
    story.append(Spacer(1, 8))

    # ── PROJECT INFO TABLE ────────────────────────────────────────────────────────
    info_data = [
        ["Project ID", project.get("project_id", ""), "Project Name", project.get("project_name", "")],
        ["Client Name", project.get("client_name", ""), "Location", project.get("location", "")],
        ["Building Type", project.get("building_type", "").title(), "Hazard Category", project.get("hazard_category", "").title()],
        ["Generated On", datetime.now().strftime("%d-%m-%Y %H:%M"), "Standard", "NBC 2016 / IS 2189"],
    ]
    info_table = Table(info_data, colWidths=[3 * cm, 5.5 * cm, 3.5 * cm, 5.5 * cm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), LIGHT_RED),
        ("BACKGROUND", (2, 0), (2, -1), LIGHT_RED),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [WHITE, LIGHT_GRAY]),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 12))

    # ── BUILDING SUMMARY ──────────────────────────────────────────────────────────
    story.append(Paragraph("BUILDING ANALYSIS SUMMARY", ParagraphStyle(
        "SectionHead", parent=styles["Heading2"],
        fontSize=11, textColor=DARK_RED, spaceAfter=4,
    )))
    summary_data = [
        ["Parameter", "Value", "Parameter", "Value"],
        ["Total Area", f"{building_data.get('estimated_area', 0):.0f} sqm",
         "Floors", str(building_data.get("floors", 1))],
        ["Rooms", str(building_data.get("rooms", 0)),
         "Corridors", str(building_data.get("corridors", 0))],
        ["Staircases", str(building_data.get("stairs", 0)),
         "Exits", str(building_data.get("exits", 0))],
    ]
    summary_table = Table(summary_data, colWidths=[4 * cm, 4.5 * cm, 4 * cm, 4.5 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DARK_GRAY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, 1), (0, -1), LIGHT_RED),
        ("BACKGROUND", (2, 1), (2, -1), LIGHT_RED),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 12))

    # ── BOQ SECTIONS ─────────────────────────────────────────────────────────────
    boq_header = ["S.No", "Item", "Description", "Unit", "Quantity", "Calculation Basis"]
    col_widths = [1 * cm, 3.5 * cm, 6.5 * cm, 1.2 * cm, 1.8 * cm, 3.5 * cm]

    section_colors = {
        "A": colors.HexColor("#E74C3C"),
        "B": colors.HexColor("#2980B9"),
        "C": colors.HexColor("#27AE60"),
    }

    for section in boq_sections:
        sec_id = section.get("section_id", "")
        sec_color = section_colors.get(sec_id, DARK_GRAY)
        sec_light = colors.HexColor(
            f"#{min(255, int(sec_color.hexval()[1:3], 16) + 40):02X}"
            f"{min(255, int(sec_color.hexval()[3:5], 16) + 40):02X}"
            f"{min(255, int(sec_color.hexval()[5:7], 16) + 40):02X}"
        )

        # Section heading
        story.append(Paragraph(
            f"SECTION {sec_id}: {section.get('section_name', '').upper()}",
            ParagraphStyle("SecHead", parent=styles["Heading2"],
                           fontSize=10, textColor=sec_color, spaceAfter=4, spaceBefore=8)
        ))

        table_data = [boq_header]
        for item in section.get("items", []):
            table_data.append([
                str(item.get("sno", "")),
                item.get("item", ""),
                item.get("description", ""),
                item.get("unit", ""),
                f"{item.get('quantity', 0):.1f}",
                item.get("calculation_basis", ""),
            ])

        boq_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        boq_table.setStyle(TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), DARK_GRAY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.5),
            # Data
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
            ("GRID", (0, 0), (-1, -1), 0.3, MID_GRAY),
            ("PADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),  # S.No
            ("ALIGN", (4, 0), (4, -1), "RIGHT"),   # Quantity
            # Item column bold
            ("FONTNAME", (1, 1), (1, -1), "Helvetica-Bold"),
        ]))
        story.append(boq_table)
        story.append(Spacer(1, 8))

    # ── FOOTER NOTE ───────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY))
    story.append(Spacer(1, 4))
    note_style = ParagraphStyle("Note", parent=styles["Normal"], fontSize=6.5,
                                textColor=DARK_GRAY, spaceAfter=2)
    story.append(Paragraph(
        "⚠ NOTE: This BOQ is generated by AI analysis and should be verified by a qualified fire safety engineer. "
        "Quantities are approximate and subject to site verification. "
        "Standards applied: NBC 2016 Part 4, IS 2189:2008, IS 15105:2002, IS 3844:1989, IS 2190:2010.",
        note_style
    ))
    story.append(Paragraph(
        f"Generated by Fire BOQ Platform | AI-Powered | {datetime.now().strftime('%d %B %Y %H:%M')}",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=6, textColor=MID_GRAY,
                       alignment=TA_CENTER)
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def generate_excel(project: dict, building_data: dict, recommendations: dict, boq_sections: list) -> bytes:
    """Generate a formatted Excel BOQ report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fire BOQ"

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35

    # Style helpers
    def header_style(cell, bg="#2C3E50", fg="FFFFFF", bold=True, size=10):
        cell.fill = PatternFill("solid", fgColor=bg.lstrip("#"))
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def data_style(cell, bold=False, align="left", size=9):
        cell.font = Font(bold=bold, size=size)
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ── TITLE ─────────────────────────────────────────────────────────────────────
    row = 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = "FIRE PROTECTION SYSTEM — BILL OF QUANTITIES"
    header_style(cell, "#C0392B", size=14)
    ws.row_dimensions[row].height = 28

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = f"Project: {project.get('project_name', '')}  |  ID: {project.get('project_id', '')}  |  Client: {project.get('client_name', '')}"
    cell.font = Font(bold=True, size=10, color="2C3E50")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 18

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = (
        f"Location: {project.get('location', '')}  |  Building Type: {project.get('building_type', '').title()}  "
        f"|  Hazard: {project.get('hazard_category', '').title()}  "
        f"|  Date: {datetime.now().strftime('%d-%m-%Y')}"
    )
    cell.font = Font(size=9, color="7F8C8D")
    cell.alignment = Alignment(horizontal="center")

    row += 2

    # ── BOQ SECTIONS ─────────────────────────────────────────────────────────────
    section_colors = {"A": "E74C3C", "B": "2980B9", "C": "27AE60"}

    for section in boq_sections:
        sec_id = section.get("section_id", "")
        sec_color = section_colors.get(sec_id, "2C3E50")

        # Section header
        ws.merge_cells(f"A{row}:F{row}")
        cell = ws[f"A{row}"]
        cell.value = f"SECTION {sec_id}: {section.get('section_name', '').upper()}"
        header_style(cell, f"#{sec_color}", size=11)
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers
        headers = ["S.No", "Item", "Description", "Unit", "Quantity", "Calculation Basis"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            header_style(cell, "#34495E", size=9)
            cell.border = thin
        ws.row_dimensions[row].height = 16
        row += 1

        # Items
        alt_colors = ["FFFFFF", "F8F9FA"]
        for i, item in enumerate(section.get("items", [])):
            bg = alt_colors[i % 2]
            row_data = [
                item.get("sno", ""),
                item.get("item", ""),
                item.get("description", ""),
                item.get("unit", ""),
                item.get("quantity", 0),
                item.get("calculation_basis", ""),
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = PatternFill("solid", fgColor=bg)
                bold = col == 2  # item name bold
                align = "right" if col == 5 else ("center" if col in [1, 4] else "left")
                data_style(cell, bold=bold, align=align)
                cell.border = thin
            ws.row_dimensions[row].height = 40
            row += 1

        row += 1  # Gap between sections

    # ── NOTES ─────────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = (
        "NOTE: BOQ generated by AI analysis. Quantities subject to site verification. "
        "Standards: NBC 2016 Part 4, IS 2189:2008, IS 15105:2002."
    )
    cell.font = Font(italic=True, size=8, color="7F8C8D")
    cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_csv(boq_sections: list) -> str:
    """Generate CSV from BOQ sections."""
    import csv
    import io as _io

    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Section", "S.No", "Item", "Description", "Unit", "Quantity", "Calculation Basis"])

    for section in boq_sections:
        sec_name = section.get("section_name", "")
        for item in section.get("items", []):
            writer.writerow([
                sec_name,
                item.get("sno", ""),
                item.get("item", ""),
                item.get("description", ""),
                item.get("unit", ""),
                item.get("quantity", 0),
                item.get("calculation_basis", ""),
            ])

    return output.getvalue()
